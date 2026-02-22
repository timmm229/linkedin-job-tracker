#!/usr/bin/env python3
"""
LinkedIn Job Alert Email Parser - ENHANCED VERSION
Fetches actual job details from LinkedIn URLs
"""

import imaplib
import email
from email.header import decode_header
import re
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import requests
import time
import json

# Email Configuration
EMAIL_ACCOUNT = os.environ.get('EMAIL_ADDRESS', '')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')
IMAP_SERVER = os.environ.get('IMAP_SERVER', 'imap.gmail.com')

EXCEL_FILE = "linkedin_jobs_tracker.xlsx"

# Load keywords from external config file
def load_keywords():
    """Load keywords from keywords.json file"""
    try:
        with open('keywords.json', 'r') as f:
            config = json.load(f)
            return config.get('HIGH_PRIORITY_KEYWORDS', []), config.get('MEDIUM_PRIORITY_KEYWORDS', [])
    except FileNotFoundError:
        print("WARNING: keywords.json not found. Creating default file...")
        create_default_keywords()
        return load_keywords()
    except json.JSONDecodeError:
        print("ERROR: keywords.json is not valid JSON. Please fix the file.")
        return [], []

def create_default_keywords():
    """Create default keywords.json file"""
    default_config = {
        "HIGH_PRIORITY_KEYWORDS": [
            "oracle erp", "oracle epm", "technical sales", "fusion", "netsuite",
            "manager", "senior manager", "pwc", "pricewaterhousecoopers"
        ],
        "MEDIUM_PRIORITY_KEYWORDS": [
            "oracle cloud", "oracle application", "oracle consultant", 
            "oracle developer", "oracle hcm", "oracle scm"
        ]
    }
    with open('keywords.json', 'w') as f:
        json.dump(default_config, f, indent=4)
    print("Created keywords.json with default keywords")

# Load keywords at startup
HIGH_PRIORITY_KEYWORDS, MEDIUM_PRIORITY_KEYWORDS = load_keywords()

def scrape_linkedin_job(url):
    """
    Scrape job details directly from LinkedIn URL
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            
            job_data = {
                'title': 'Not found',
                'company': 'Not found',
                'location': 'Not found',
                'url': url
            }
            
            # Extract title - multiple methods
            title_elem = soup.find('h1', class_=re.compile('top-card-layout__title')) or \
                        soup.find('h1') or \
                        soup.find('title')
            
            if title_elem:
                title_text = title_elem.get_text(strip=True)
                # Clean up title (remove " - LinkedIn" etc)
                title_text = re.sub(r'\s*-\s*LinkedIn.*$', '', title_text)
                title_text = re.sub(r'\s*\|\s*LinkedIn.*$', '', title_text)
                job_data['title'] = title_text[:200] if title_text else 'Not found'
            
            # Extract company
            company_elem = soup.find('a', class_=re.compile('topcard__org-name-link')) or \
                          soup.find('span', class_=re.compile('topcard__flavor')) or \
                          soup.find('a', href=re.compile('/company/'))
            
            if company_elem:
                company_text = company_elem.get_text(strip=True)
                job_data['company'] = company_text[:100] if company_text else 'Not found'
            
            # Extract location
            location_elem = soup.find('span', class_=re.compile('topcard__flavor--bullet')) or \
                           soup.find(text=re.compile(r'[A-Z][a-z]+,\s*[A-Z]{2}'))
            
            if location_elem:
                if hasattr(location_elem, 'get_text'):
                    location_text = location_elem.get_text(strip=True)
                else:
                    location_text = str(location_elem).strip()
                job_data['location'] = location_text[:100] if location_text else 'Not found'
            
            print(f"    Scraped: {job_data['title'][:50]} at {job_data['company'][:30]}")
            return job_data
            
        else:
            print(f"    Failed to fetch URL (status {response.status_code})")
            return None
            
    except Exception as e:
        print(f"    Error scraping URL: {str(e)}")
        return None

def connect_to_email():
    """Connect to email account via IMAP"""
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
        return mail
    except Exception as e:
        print(f"Error connecting to email: {str(e)}")
        return None

def extract_job_urls_from_email(email_body):
    """Extract all LinkedIn job URLs from email"""
    url_patterns = [
        r'https://(?:www\.)?linkedin\.com/jobs/view/(\d+)',
        r'https://(?:www\.)?linkedin\.com/comm/jobs/view/(\d+)',
    ]
    
    urls = set()
    for pattern in url_patterns:
        matches = re.findall(pattern, email_body, re.IGNORECASE)
        for match in matches:
            job_id = match
            full_url = f"https://www.linkedin.com/jobs/view/{job_id}"
            urls.add(full_url)
    
    return list(urls)

def fetch_linkedin_job_emails(mail, days_back=30):
    """Fetch LinkedIn job alert emails from the last N days"""
    try:
        mail.select('INBOX')
        search_criteria = '(FROM "linkedin.com")'
        status, messages = mail.search(None, search_criteria)
        
        if status != 'OK':
            print("No emails found")
            return []
        
        email_ids = messages[0].split()
        print(f"Found {len(email_ids)} LinkedIn emails total")
        
        recent_emails = email_ids[-50:] if len(email_ids) > 50 else email_ids
        print(f"Processing last {len(recent_emails)} emails...\n")
        
        all_urls = []
        
        for email_id in recent_emails:
            try:
                status, msg_data = mail.fetch(email_id, '(RFC822)')
                if status != 'OK':
                    continue
                
                email_body = msg_data[0][1]
                email_message = email.message_from_bytes(email_body)
                
                # Get email date
                date_tuple = email.utils.parsedate_tz(email_message['Date'])
                if date_tuple:
                    email_date = datetime.fromtimestamp(email.utils.mktime_tz(date_tuple))
                    
                    # Only 2026 emails
                    if email_date.year != 2026:
                        continue
                    
                    # Skip if older than days_back
                    if (datetime.now() - email_date).days > days_back:
                        continue
                
                # Extract body
                body = get_email_body(email_message)
                
                if body:
                    urls = extract_job_urls_from_email(body)
                    all_urls.extend(urls)
                    
            except Exception as e:
                print(f"Error processing email: {str(e)}")
                continue
        
        print(f"\nTotal unique job URLs found: {len(set(all_urls))}")
        return list(set(all_urls))
        
    except Exception as e:
        print(f"Error fetching emails: {str(e)}")
        return []

def get_email_body(email_message):
    """Extract email body from email message"""
    body = ""
    
    if email_message.is_multipart():
        for part in email_message.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))
            
            if content_type == "text/html" and "attachment" not in content_disposition:
                try:
                    body = part.get_payload(decode=True).decode()
                    break
                except:
                    pass
    else:
        try:
            body = email_message.get_payload(decode=True).decode()
        except:
            pass
    
    return body

def calculate_priority(job_data):
    """Calculate priority score"""
    title = job_data.get('title', '').lower()
    company = job_data.get('company', '').lower()
    combined_text = f"{title} {company}".lower()
    
    for keyword in HIGH_PRIORITY_KEYWORDS:
        if keyword in combined_text:
            return 1
    
    for keyword in MEDIUM_PRIORITY_KEYWORDS:
        if keyword in combined_text:
            return 2
    
    return 3

def initialize_spreadsheet():
    """Create new Excel file with headers"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Job Postings"
    
    headers = [
        'Priority', 'Job Title', 'Company', 'Location', 
        'Travel Required', 'Salary/Rate', 'Job URL', 'Date Added'
    ]
    
    sheet.append(headers)
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    column_widths = [10, 50, 30, 30, 15, 15, 50, 15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    
    sheet.freeze_panes = 'A2'
    wb.save(EXCEL_FILE)
    print(f"Created new spreadsheet: {EXCEL_FILE}")

def add_jobs_to_spreadsheet(job_urls):
    """Scrape job details and add to spreadsheet"""
    if not os.path.exists(EXCEL_FILE):
        initialize_spreadsheet()
    
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    # Get existing URLs
    existing_urls = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[6]:
            existing_urls.add(row[6])
    
    print(f"\nExisting jobs in spreadsheet: {len(existing_urls)}")
    print(f"\nScraping job details from LinkedIn...\n")
    
    new_jobs = []
    
    for i, url in enumerate(job_urls, 1):
        if url in existing_urls:
            print(f"{i}/{len(job_urls)}: SKIP (already exists) {url[:60]}")
            continue
        
        print(f"{i}/{len(job_urls)}: Fetching {url[:60]}")
        
        job_data = scrape_linkedin_job(url)
        
        if job_data:
            # Skip jobs with "Not specified" location
            if job_data.get('location', '').lower() == 'not specified':
                print(f"    SKIPPED: Location is 'Not specified'")
                continue
            
            job_data['priority'] = calculate_priority(job_data)
            new_jobs.append(job_data)
            time.sleep(2)  # Be respectful to LinkedIn servers
    
    # Sort by priority
    new_jobs.sort(key=lambda x: x['priority'])
    
    # Add to spreadsheet
    current_date = datetime.now().strftime('%Y-%m-%d')
    new_count = 0
    
    for job in new_jobs:
        row_data = [
            job['priority'],
            job.get('title', 'Not found'),
            job.get('company', 'Not found'),
            job.get('location', 'Not found'),
            'Not specified',
            'Not Listed',
            job.get('url', ''),
            current_date
        ]
        
        sheet.append(row_data)
        new_count += 1
        
        # Color coding
        row_num = sheet.max_row
        priority_cell = sheet.cell(row=row_num, column=1)
        
        if job['priority'] == 1:
            priority_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            priority_cell.font = Font(bold=True)
        elif job['priority'] == 2:
            priority_cell.fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    
    wb.save(EXCEL_FILE)
    print(f"\n[SUCCESS] Added {new_count} new jobs to spreadsheet")
    return new_count

def main():
    """Main execution"""
    print(f"LinkedIn Job Alert Email Parser - ENHANCED")
    print(f"{'='*70}")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    if not EMAIL_ACCOUNT or not EMAIL_PASSWORD:
        print("ERROR: Email credentials not set")
        return
    
    # Connect to email
    print(f"Connecting to {IMAP_SERVER}...")
    mail = connect_to_email()
    
    if not mail:
        return
    
    print("Connected!\n")
    
    # Fetch job URLs from emails
    job_urls = fetch_linkedin_job_emails(mail, days_back=30)
    mail.logout()
    
    if not job_urls:
        print("No job URLs found in emails")
        return
    
    # Scrape and add jobs
    new_count = add_jobs_to_spreadsheet(job_urls)
    
    print(f"\n{'='*70}")
    print(f"Summary:")
    print(f"- Job URLs found: {len(job_urls)}")
    print(f"- New jobs added: {new_count}")
    print(f"- Spreadsheet: {EXCEL_FILE}")
    print(f"\nPriority 1 jobs (Green) are at the top!")

    input("Press Enter to exit...")

if __name__ == "__main__":
    main()
